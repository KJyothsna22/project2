"""
Automated Reporting Engine
Generates:
1. CSV Data Reports (Forecasting & Inventory Recommendations)
2. Professional Multi-Tab Excel (.xlsx) Reports with custom styling and KPI summaries
3. Printable Executive Briefing PDF / HTML Reports (Executive Summary, Stockout Alerts, Purchase Orders)
"""

import os
from datetime import datetime
from io import BytesIO
from typing import Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
import config
from warehouse.db import WarehouseManager
from utils.logger import get_logger

logger = get_logger("report_generator")

class PDFReport(FPDF):
    """Custom PDF Generator for Executive Briefings."""

    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(24, 43, 73)
        self.cell(0, 10, "RETAIL DEMAND FORECASTING & INVENTORY OPTIMIZATION PLATFORM", border=False, ln=True, align="L")
        self.set_font("Helvetica", "I", 9)
        self.set_text_color(100, 116, 139)
        self.cell(0, 6, f"Executive Briefing & Inventory Health Audit | Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", border=False, ln=True, align="L")
        self.ln(3)
        self.set_draw_color(203, 213, 225)
        self.set_line_width(0.5)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(148, 163, 184)
        self.cell(0, 10, f"Page {self.page_no()} | Confidential - Enterprise Supply Chain Operations", border=False, align="C")

class ReportGenerator:
    """Generates CSV, Excel, and PDF reports."""

    @classmethod
    def generate_excel_report(cls) -> BytesIO:
        """
        Creates a multi-tab styled Excel workbook containing:
        - Sheet 1: Executive KPI Summary
        - Sheet 2: Inventory Optimization & Risk Alerts
        - Sheet 3: Demand Forecasts
        """
        logger.info("Generating multi-tab styled Excel report...")
        output = BytesIO()

        # Fetch data
        inv_df = WarehouseManager.read_query("SELECT * FROM fact_inventory_recommendations")
        fc_df = WarehouseManager.read_query("SELECT * FROM fact_forecast_results WHERE actual_demand IS NULL LIMIT 200")
        store_df = WarehouseManager.read_query("SELECT * FROM mart_store_performance")

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # 1. Executive Summary Sheet
            summary_stats = {
                "Metric": [
                    "Total Monitored SKUs",
                    "Total On-Hand Inventory Units",
                    "Total Forecast Replenishment Units",
                    "Estimated Replenishment Cost",
                    "Critical Stockout Alert Count",
                    "Excess Overstock Alert Count",
                    "Stores Operating",
                    "Reporting Date"
                ],
                "Value": [
                    len(inv_df),
                    f"{inv_df['current_inventory'].sum():,}",
                    f"{inv_df['recommended_order_qty'].sum():,}",
                    f"${inv_df['estimated_order_cost'].sum():,.2f}",
                    int((inv_df['inventory_status'] == 'Critical Stockout Risk').sum()),
                    int((inv_df['inventory_status'] == 'Excess Overstock').sum()),
                    len(inv_df['store_id'].unique()),
                    datetime.now().strftime("%Y-%m-%d")
                ]
            }
            pd.DataFrame(summary_stats).to_excel(writer, sheet_name="Executive_Summary", index=False)
            
            # 2. Inventory Recommendations Sheet
            inv_cols = [
                "item_id", "store_id", "current_inventory", "average_daily_demand",
                "safety_stock", "reorder_point", "recommended_order_qty",
                "days_of_supply", "inventory_status", "risk_level",
                "stockout_risk_pct", "suggested_replenishment_date", "estimated_order_cost"
            ]
            inv_df[[c for c in inv_cols if c in inv_df.columns]].to_excel(
                writer, sheet_name="Inventory_Planning", index=False
            )

            # 3. Forecasts Sheet
            fc_cols = ["forecast_date", "item_id", "store_id", "model_used", "forecast_demand", "confidence_lower", "confidence_upper", "horizon_days"]
            fc_df[[c for c in fc_cols if c in fc_df.columns]].to_excel(
                writer, sheet_name="Demand_Forecasts", index=False
            )

            # 4. Store Performance Sheet
            store_df.to_excel(writer, sheet_name="Store_Performance", index=False)

        output.seek(0)
        return output

    @classmethod
    def generate_pdf_report(cls) -> bytes:
        """
        Creates an executive PDF briefing report with KPI cards, risk breakdown,
        and priority replenishment action items.
        """
        logger.info("Generating Executive Briefing PDF report...")
        pdf = PDFReport()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        # 1. Fetch Summary Data
        inv_df = WarehouseManager.read_query("SELECT * FROM fact_inventory_recommendations")
        total_skus = len(inv_df)
        total_inv = int(inv_df["current_inventory"].sum())
        total_reorder = int(inv_df["recommended_order_qty"].sum())
        total_cost = float(inv_df["estimated_order_cost"].sum())
        high_risk_count = int((inv_df["risk_level"] == "High Risk").sum())
        stockouts = int((inv_df["inventory_status"] == "Critical Stockout Risk").sum())

        # Section 1: Executive KPI Highlights
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(0, 8, "1. Executive Supply Chain Summary", ln=True)
        pdf.ln(2)

        pdf.set_font("Helvetica", "", 10)
        pdf.set_fill_color(241, 245, 249)
        pdf.rect(10, pdf.get_y(), 190, 24, "F")
        
        pdf.set_xy(12, pdf.get_y() + 3)
        pdf.cell(60, 6, f"Total Monitored SKUs: {total_skus}", ln=False)
        pdf.cell(65, 6, f"On-Hand Inventory: {total_inv:,} units", ln=False)
        pdf.cell(65, 6, f"High Risk SKUs: {high_risk_count} ({high_risk_count/max(1,total_skus)*100:.1f}%)", ln=True)
        
        pdf.set_x(12)
        pdf.cell(60, 6, f"Recommended Order: {total_reorder:,} units", ln=False)
        pdf.cell(65, 6, f"Est. Replenishment Cost: ${total_cost:,.2f}", ln=False)
        pdf.cell(65, 6, f"Critical Stockouts: {stockouts}", ln=True)
        pdf.ln(8)

        # Section 2: Critical Stockout & Replenishment Alerts
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "2. Priority Inventory Action Items (Top High-Risk SKUs)", ln=True)
        pdf.ln(2)

        # Table Header
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(255, 255, 255)
        
        pdf.cell(30, 7, "Item ID", border=1, fill=True)
        pdf.cell(20, 7, "Store", border=1, fill=True)
        pdf.cell(22, 7, "On-Hand", border=1, fill=True, align="R")
        pdf.cell(22, 7, "Safety Stk", border=1, fill=True, align="R")
        pdf.cell(22, 7, "Reorder Pt", border=1, fill=True, align="R")
        pdf.cell(25, 7, "Order Qty", border=1, fill=True, align="R")
        pdf.cell(25, 7, "Target Date", border=1, fill=True, align="C")
        pdf.cell(24, 7, "Status", border=1, fill=True, align="C")
        pdf.ln()

        # Table Rows (Top 12 high risk items)
        high_risk_df = inv_df[inv_df["risk_level"] == "High Risk"].sort_values(by="stockout_risk_pct", ascending=False).head(12)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(30, 41, 59)
        
        for idx, row in high_risk_df.iterrows():
            fill_bg = (254, 242, 242) if row["inventory_status"] == "Critical Stockout Risk" else (255, 255, 255)
            pdf.set_fill_color(*fill_bg)
            
            pdf.cell(30, 6, str(row["item_id"]), border=1, fill=True)
            pdf.cell(20, 6, str(row["store_id"]), border=1, fill=True)
            pdf.cell(22, 6, str(row["current_inventory"]), border=1, fill=True, align="R")
            pdf.cell(22, 6, str(row["safety_stock"]), border=1, fill=True, align="R")
            pdf.cell(22, 6, str(row["reorder_point"]), border=1, fill=True, align="R")
            pdf.cell(25, 6, str(row["recommended_order_qty"]), border=1, fill=True, align="R")
            pdf.cell(25, 6, str(row["suggested_replenishment_date"]), border=1, fill=True, align="C")
            pdf.cell(24, 6, "Stockout" if row["inventory_status"] == "Critical Stockout Risk" else "Overstock", border=1, fill=True, align="C")
            pdf.ln()

        pdf.ln(8)
        # Sign-off block
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 6, "Supply Chain VP Sign-off: _______________________      Date: ______________", ln=True)

        return bytes(pdf.output())

    @classmethod
    def generate_csv_report(cls, table_name: str = "fact_inventory_recommendations") -> str:
        """Returns CSV string for a given table."""
        df = WarehouseManager.read_query(f"SELECT * FROM {table_name}")
        return df.to_csv(index=False)
